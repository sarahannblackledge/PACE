import numpy as np
import pydicom as dicom
import os
import SimpleITK as sitk
import sys
import time
from openpyxl import Workbook, load_workbook
sys.path.append('/Users/sblackledge/PycharmProjects/PACE/PACE')
from dicom_utilities import copy_dicom_tags
from key import key
from create_rtstruct_masks_SB2 import create_rtstruct_masks

floc_el = 0x19100c  # Used to store the file location in read dicoms
pat_directory = '/Users/sblackledge/DATA/ProKnow_database/RMH_proknow/proknowPACE'
dose_directory = os.path.join(pat_directory, "RTDOSE")
rtstruct_directory = os.path.join(pat_directory, "RTSTRUCT")
plan_directory = os.path.join(pat_directory, "RTPLAN")
mr_directory = os.path.join(pat_directory, "MR")
output_directory = os.path.join(pat_directory, "nifti_dump4")
masks_of_interest = ['ProstateOnly', 'SVsOnly', 'Rectum', 'Bladder', 'Bowel', 'PenileBulb', 'Penile_Bulb', 'External']


# Functions to build master dictionaries/lists containing metadata from every dicom file in specified directory
def build_mr_dict(mr_directory, floc_el):
    """Loops through subfolders in dicom dump folder to store metadata of each image using the series UID as
    the dictionary key.
    inputs:
        1. mr_directory - str: full path to the directory where the subfolders containing the MR dicoms from each
        individual scan are stored.
    output:
        1. dict_mr_dicoms - dict: Each value in the dictionary is a list (sorted by slice location) of the metadata for
        every dicom file comprising a single scan. The dictionary key is the dicom series UID for each scan.
        """
    # Start timer
    start_time = time.time()

    study_uids_blacklist = {}

    # Load in the mr dicoms
    dict_mr_dicoms = {}
    for dcm_folder in os.listdir(mr_directory):
        if dcm_folder == ".DS_Store":
            continue
        dirpath_scan = os.path.join(mr_directory, dcm_folder)
        for file in os.listdir(dirpath_scan):
            try:
                dicom_path = os.path.join(dirpath_scan, file)
                dcm_obj = dicom.read_file(dicom_path, stop_before_pixels=True)
                series_uid = dcm_obj.SeriesInstanceUID
                if dcm_obj.StudyInstanceUID in study_uids_blacklist.keys():
                    break
                if not series_uid in dict_mr_dicoms:
                    dict_mr_dicoms[series_uid] = []

                dcm_obj.add_new(floc_el, "ST", dicom_path)
                dict_mr_dicoms[series_uid].append(dcm_obj)
            except:
                raise

        # Now organise by ascending slice location
        for series_uid in dict_mr_dicoms:
            slice_locations = [float(dcm.ImagePositionPatient[-1]) for dcm in dict_mr_dicoms[series_uid]]
            dict_mr_dicoms[series_uid] = np.array(dict_mr_dicoms[series_uid])[np.argsort(slice_locations)].tolist()

    # End timer
    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f"Elapsed time: {elapsed_time:.4f} seconds")

    return dict_mr_dicoms


def build_dcm_list(dcm_directory, floc_el):
    """Loops through files in a dicom directory to build a list where each item contains the dicom metadata for
    each individual dicom file. This works for reading in both RTSTRUCTS and RTPLANS
    inputs:
        1. dcm_directory - str: full path to directory where RTSTRUCT or RTPLANS are stored.
    output:
        1. dicom_metadata_list - list: List of dicom metadata (which is a list itself) for each file in the dcm_directory
        folder. The items in the list are numbered according to the order in which they were read."""

    study_uids_blacklist = {}

    dicom_metadata_list = []
    for dicom_file in os.listdir(dcm_directory):
        if dicom_file == ".DS_Store":
            continue
        try:
            dcm_path = os.path.join(dcm_directory, dicom_file)
            dcm = dicom.read_file(dcm_path)
            if dcm.StudyInstanceUID in study_uids_blacklist.keys():
                continue
            dcm.add_new(floc_el, "ST", dcm_path)
            dicom_metadata_list.append(dcm)
        except:
            raise

    print(f"List built for {dcm_directory}")
    return dicom_metadata_list


def build_dose_list(dose_directory, floc_el):
    """Loops through files in a dose cube directory to build a list where each item contains the dicom metadata for
       each individual RTDOSE file.
       inputs:
           1. dose_directory - str: full path to directory where RTDOSE files are stored.
       output:
           1. dose_dicoms_list - list: List of dicom metadata (which is a list itself) for each RTDOSE dicom file in the
            dcm_directory folder. The items in the list are numbered according to the order in which they were read."""

    study_uids_blacklist = {}

    # Initialize list
    dose_dicoms_list = []
    for dicom_file in os.listdir(dose_directory):
        if dicom_file == ".DS_Store":
            continue
        try:
            dose_path = os.path.join(dose_directory, dicom_file)
            dcm = dicom.read_file(dose_path)

            if dcm.StudyInstanceUID in study_uids_blacklist.keys():
                continue
            if not dcm.DoseSummationType == "PLAN":
                continue  # Ignore beams and fractions.

            add_dose = True
            # Check if dose cube already in list; if so, difference between dose cubes will be ~0
            for dose_dicom in dose_dicoms_list:
                dose_array_new = dcm.pixel_array
                dose_array_old = dose_dicom.pixel_array
                if np.allclose(dose_array_new.shape, dose_array_old.shape):
                    if np.abs(np.sum(dose_array_old - dose_array_new)) < 1e-5:
                        add_dose = False
            if add_dose:
                dcm.add_new(floc_el, "ST", dose_path)
                dose_dicoms_list.append(dcm)
        except:
            raise

    print(f"List built for {dose_directory}")


    return dose_dicoms_list


# Functions to find corresponding dicom files based on metadata tags
def find_plan_from_dose(dose_dicom, plan_dicoms):
    ref_plan_uid = dose_dicom[0x300c, 0x0002][0][0x0008, 0x1155].value
    ref_plan = None
    for plan in plan_dicoms:
        if plan.SOPInstanceUID == ref_plan_uid:
            ref_plan = plan
            db_id = ref_plan.StudyDescription
            plan_fname = os.path.basename(plan.filename)

    if ref_plan is None:
        print("Could not find a plan for dose cube: %s" % ref_plan_uid)
        plan_fname = 'None'
        db_id = 'None'

    return ref_plan, db_id, plan_fname


def find_rtstruct_from_plan(ref_plan, rtstruct_dicoms):
    ref_rtstruct_uid = ref_plan[0x300c, 0x0060][0][0x0008, 0x1155].value
    ref_rtstruct_uid_str = str(ref_rtstruct_uid)
    ref_rtstruct = None
    for rtstruct in rtstruct_dicoms:
        if rtstruct.SOPInstanceUID == ref_rtstruct_uid:
            ref_rtstruct = rtstruct
            rtstruct_fname = os.path.basename(ref_rtstruct.filename)
            rtstruct_id = ref_rtstruct.StudyDescription

    if ref_rtstruct is None:
        print("Could not find a rtstruct for dose cube: %s" % ref_rtstruct_uid_str)
        rtstruct_fname = 'None'
        rtstruct_id = 'None'

    return ref_rtstruct, rtstruct_fname, rtstruct_id


def find_mr_from_rtstruct(ref_rtstruct, mr_dicoms):
    ref_mr_series_uid = ref_rtstruct[0x3006, 0x10][0][0x3006, 0x12][0][0x3006, 0x14][0][0x20, 0xe].value
    ref_mr_series_uid_str = str(ref_mr_series_uid)
    for series_uid in mr_dicoms:
        if series_uid == ref_mr_series_uid:
            ref_mr_study = mr_dicoms[series_uid]
            mr_dirname = str(series_uid)
            mr_study_description = ref_mr_study[0].StudyDescription

    if ref_mr_study is None:
        print("Could not find an MR series for dose cube: %s" % ref_mr_series_uid_str)
        mr_dirname = 'None'

    return ref_mr_study, mr_dirname, mr_study_description


# Function to find dicom id (as encoded in MR image 'Study Description') in key.py file
def databaseID_to_patientID(mr_study_description, i):
    # Read in data from key.py (hard-coded)
    id_arr = key()[:, 1]

    index = np.where(id_arr == mr_study_description)[0]
    TF = np.any(index)

    if TF:
        index = index[0]
        names_arr = key()[:, 0]
        pat_id = names_arr[index]
    else:
        print(f"MR study description {mr_study_description} does not match any entry in db_id key")
        pat_id = 'unknown_%d' % i
        db_id = str(i)

    return pat_id


# Functions to write to .nii.gz
def write_mr(output_directory, fname, ref_mr_study, floc_el):
    path_mri = os.path.join(output_directory, 'MRI')

    if not os.path.exists(path_mri):
        os.makedirs(path_mri)

    # Create the MR image if not already there
    study_path_mri = os.path.join(path_mri, fname)
    if os.path.exists(study_path_mri):
        mr_sitk = sitk.ReadImage(study_path_mri)
    else:
        mr_sitk = sitk.ReadImage([dcm[floc_el].value for dcm in ref_mr_study])
        copy_dicom_tags(mr_sitk, ref_mr_study[0], ignore_private=True)
        sitk.WriteImage(mr_sitk, study_path_mri, True)
    print(f'MR image saved to {study_path_mri}')

    return mr_sitk


def write_dose(output_directory, dose_dicom, fname, mr_sitk, floc_el):
    path_dose = os.path.join(output_directory, "dose")
    if not os.path.exists(path_dose):
        os.makedirs(path_dose)
        print('dose subfolder generated')

    # Create the dose cube and resample to the MR image
    study_path_dose = os.path.join(path_dose, fname)

    scaling = dose_dicom[0x3004, 0xe].value
    dose_sitk = sitk.ReadImage(dose_dicom[floc_el].value)
    dose_sitk = sitk.Resample(dose_sitk, mr_sitk, sitk.AffineTransform(3), sitk.sitkLinear, 0.0,
                                      sitk.sitkFloat64)
    dose_sitk = dose_sitk * scaling
    copy_dicom_tags(dose_sitk, dose_dicom, ignore_private=True)
    sitk.WriteImage(dose_sitk, study_path_dose, True)

    print(f'Dose cube saved to {study_path_dose}')

    return dose_sitk

def write_masks(output_directory, ref_rtstruct, mr_sitk, fname, masks_of_interest):
    path_masks3D = os.path.join(output_directory, 'masks3D')
    if not os.path.exists(path_masks3D):
        os.makedirs(path_masks3D)

    # Convert RTSTRUCT to sitk masks for each desired structure
    rtstruct_images_sub = create_rtstruct_masks(ref_rtstruct, mr_sitk, masks_of_interest)

    for im in rtstruct_images_sub:
        organ = im.GetMetaData("ContourName")
        fpath_organ = os.path.join(path_masks3D, organ)
        savename = os.path.join(fpath_organ, fname)
        if not os.path.exists(fpath_organ):
            os.makedirs(fpath_organ)

        sitk.WriteImage(im, savename, True)
        print(f'Mask saved to {savename}')

    return rtstruct_images_sub


def dicom_convert(dose_dicoms, plan_dicoms, rtstruct_dicoms, mr_dicoms, output_directory, masks_of_interest):
    # Initialize lists and filename for storing data for retrospective use (if need to find corresponding raw dicoms)
    fname_excel = 'data_log.xlsx'
    fpath_excel = os.path.join(output_directory, fname_excel)

    data_log = []
    headers = ['Study Description', 'Dose fname', 'Plan fname', 'RTSTRUCT fname', 'MR fname']

    # Update excel file
    try:
        workbook = load_workbook(fpath_excel)  # Load of exists
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()  # Create a new workbook if file doesn't exist
        sheet = workbook.active
        sheet.append(headers)

    # Loop through dose_dicoms to find corresponding plans, rtstructs, and MR images
    for counter, dose_dicom in enumerate(dose_dicoms):
        dose_fname = os.path.basename(dose_dicom.filename)

        # Find the plan corresponding to dose cube (plan UID encoded in dose cube metadata)
        ref_plan, dose_id, plan_fname = find_plan_from_dose(dose_dicom, plan_dicoms)

        # Find corresponding contour set (RTSTRUCT uid encoded in plan dicom metadata)
        ref_rtstruct, rtstruct_fname, rtstruct_id = find_rtstruct_from_plan(ref_plan, rtstruct_dicoms)
        if rtstruct_id == 'None':
            print('skipping this dose')
            continue

        # Find corresponding MR scan (MR uid encoded in rtstruct dicom metadata)
        ref_mr_study, mr_dirname, mr_study_description = find_mr_from_rtstruct(ref_rtstruct, mr_dicoms)

        # compile corresponding dicom filenames into list
        scan_row = [mr_study_description, dose_fname, plan_fname, rtstruct_fname, mr_dirname]
        data_log.append(scan_row)

        # Look up patient id based on database ID number. Hard coded in 'db_id_key.py'
        pat_id = databaseID_to_patientID(mr_study_description, counter)
        fname = pat_id + "_MR" + mr_study_description + ".nii.gz"

        # Now save nifti files!
        mr_image = write_mr(output_directory, fname, ref_mr_study, floc_el)
        dose_image = write_dose(output_directory, dose_dicom, fname, mr_image, floc_el)
        rtstruct_images = write_masks(output_directory, ref_rtstruct, mr_image, fname, masks_of_interest)

        # Update excel file
        sheet.append(scan_row)
        workbook.save(fpath_excel)
        print('excel file updated')

###############################################################ß
#RUN
mr_dicoms = build_mr_dict(mr_directory, floc_el)
plan_dicoms = build_dcm_list(plan_directory, floc_el)
rtstruct_dicoms = build_dcm_list(rtstruct_directory, floc_el)
dose_dicoms = build_dose_list(dose_directory, floc_el)

dicom_convert(dose_dicoms, plan_dicoms, rtstruct_dicoms, mr_dicoms, output_directory, masks_of_interest)