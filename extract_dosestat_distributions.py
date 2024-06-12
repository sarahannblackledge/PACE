import os
import openpyxl

def extract_dose_stat(fpath_source, fraction, plan_id):
    '''inputs:
        1. fpath_source - str: source Excel file where dose stats saved (e.g. PAC23004_dose_stats_final.xlsx
        2. fraction - int: Number indicating which fraction
        3. plan_id - str: Either "Session", "Verif" or "Previous". Corresponds to Excel sheet names in destination
        Excel file.'''

    #Excel file where consolidated dose stat data will be stored
    fpath_dest = '/Users/sblackledge/Documents/audit_evolutivePOTD/consolidated_dose_stats.xlsx'

    #Load desired excel file
    wb_source = openpyxl.load_workbook(fpath_source)
    wb_dest = openpyxl.load_workbook(fpath_dest)

    #Contstruct sheet name (source)from fraction
    sheet_name_source = 'Fraction ' + str(fraction)

    #Acess the desired sheets
    sheet_source = wb_source[sheet_name_source]
    sheet_dest = wb_dest[plan_id]

    #Determine cell range on source file where data should be extracted
    if plan_id == "Session":
        start_cell = 'C3'
        end_cell = 'C12'
    elif plan_id == "Verif":
        start_cell = 'C17'
        end_cell = 'C26'

    #Initialize empty list to store cell values
    dose_vals = []

    for row in sheet_source[start_cell:end_cell]:
        for cell in row:
            dose_vals.append(cell.value)

    #Define Treatment ID (patient name + fraction) and append to dose_vals list
    fname = os.path.basename(fpath_source)
    patient_name = fname.split('_')[0]
    Treatment_ID = [patient_name + '_Tx' + str(fraction)]
    new_row = Treatment_ID + dose_vals

    #See whether treatment ID already exists in destination excel file
    id_values = []

    for row in sheet_dest.iter_rows(min_row=1, max_row=sheet_dest.max_row, min_col=1, max_col=1):
        for cell in row:
            id_values.append(cell.value)

    if Treatment_ID[0] in id_values:
        index = id_values.index(Treatment_ID[0])
        row_start = index + 1
    else:
        row_start = len(id_values) + 1

    #Write to excel
    start_cell2 = ('A', row_start)
    current_column = start_cell2[0]
    for value in new_row:
        sheet_dest[current_column + str(start_cell2[1])].value = value
        current_column = chr(ord(current_column) + 1) #Move to next column

    wb_dest.save(fpath_dest)
    print('Excel file updated')

    return
#######################################################################################################
#Single fraction
fpath_root = '/Users/sblackledge/Documents/audit_evolutivePOTD'
patient_name = 'PAC23004'
fname = patient_name + '_dose_stats_final.xlsx'
fpath_source = os.path.join(fpath_root, patient_name, fname)
plan_id = 'Session'
fraction = 2


extract_dose_stat(fpath_source, fraction, plan_id)

########################################################################
#Loop through specified patients and fractions
Patient_dict = {
    'PAC2007': [1, 2, 3, 4],
    'PAC22003': [1, 2, 3, 4, 5],
    'PAC22010': [1, 2, 3, 4, 5],
    'PAC23004': [1, 2, 3, 4, 5],
    'PER022': [1, 2, 3, 4, 5],
    'PER042': [1, 2, 3, 4, 5],
    'PER085': [1, 2, 3, 4, 5],
    'PER093': [1, 2, 3, 4, 5],
    'PER110': [1, 2, 3, 4, 5]
}

fpath_root = '/Users/sblackledge/Documents/audit_evolutivePOTD'
plan_ids = ['Session', 'Verif']

for plan_id in plan_ids:
    for patient_name, fractions in Patient_dict.items():
        fname = patient_name + '_dose_stats_final.xlsx'
        fpath_source = os.path.join(fpath_root, patient_name, fname)
        for fraction in fractions:
            extract_dose_stat(fpath_source, fraction, plan_id)

