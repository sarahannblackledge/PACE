import tabula
import os
import shutil
import numpy as np
import csv
import pandas as pd
import openpyxl
import re

def typo_correct(string, char):
    #Looks for cases with two characters (i.e. underscores) and overwrites as one
    pattern = char + '{2,}'
    string = re.sub(pattern, char, string)
    return string
def fix_csv_shifts(output_path):
    df = pd.read_csv(output_path)
    last_row = len(df)-1
    #Check whether there are any non-NaN values in last column
    TF = df['TEST'].notna().any()
    counter = 0

    while TF:
        print(f'counter = {counter}')
        counter = counter + 1

        start_row = df['TEST'].first_valid_index() #Find index of first value jutting out into last row
        last_nan_col_name = df.iloc[start_row][::-1].isna().idxmax() #Find corresponding column where row segment should be shifted
        if start_row < last_row:
            try:
                end_row = df[last_nan_col_name].iloc[start_row:].first_valid_index()-1 #last row number that needs to be shifted
            except:
                end_row = last_row
        else:
            end_row = start_row

        start_col_idx = df.columns.get_loc(last_nan_col_name)

        # Shift the subsection of the DataFrame to the left by one column
        df.iloc[start_row:end_row + 1, start_col_idx:-1] = df.iloc[start_row:end_row + 1, start_col_idx + 1:].values

        # Optionally, set the last column of the shifted subsection to NaN
        df.iloc[start_row:end_row + 1, -1] = float('nan')

        TF = df['TEST'].notna().any()

    #Delete last column of dataframe
    df = df.drop(df.columns[-1], axis=1)

    # Overwrite csv file with correctly aligned dataframe
    df.to_csv(output_path, index=False)

def pdf_to_csv(fraction_name, patient_dir):

    ext1 = '.pdf'
    ext2 = '.csv'

    input_path = os.path.join(patient_dir, fraction_name + ext1)
    output_path = os.path.join(patient_dir, fraction_name + '_temp' + ext2) #temporary csv file
    fpath = os.path.join(patient_dir, fraction_name + '.xlsx') #Final excel file

    #Convert pdf into csv using tabula
    tabula.convert_into(input_path, output_path, output_format='csv', pages='all')

    #Open csv file using built-in python reader
    with open(output_path, 'r') as f:
        reader = csv.reader(f)
        data = list(reader)
        data_array = []
        #Fix hanging characters


    try: #Will only work of no weird data misalignment in csv file
        data_array = np.array(data)
        print('no data misalignment. Proceeding with other formatting corrections')
    except:
        # Extract existing headers
        headers = data[0]
        # Check if there is any data in the 13th column
        if any(len(row) > 12 and row[12] for row in data[1:]):
            print('data in 13th column. Fixing alignment')
            # Add column label to 13th column if necessary
            if len(headers) == 12:
                headers.append('TEST')

        # Write the updated CSV file with the new column heading
        with open(output_path, 'w', newline='') as outfile:
            writer = csv.writer(outfile)
            writer.writerows([headers] + data[1:])

        #Open as pandas dataframe and correct shifts
        fix_csv_shifts(output_path)

        with open(output_path, 'r') as f2:
            reader = csv.reader(f2)
            data2 = list(reader)
            data_array = []
            data_array = np.array(data2)
            print('Misalignment successfully corrected')


    structures = data_array[:, 0]

    #Fix hanging characters
    #rule: if number of characters in string < 4, then should be appended to previous line
    str_lengths = np.char.str_len(structures)
    inds = np.where(np.logical_and(str_lengths > 0, str_lengths < 4))
    inds = inds[0]
    new_names = np.char.add(structures[inds-1], structures[inds])
    structures[inds-1] = new_names
    structures[inds] = ''

    #Add structure name to every line
    for i, name in enumerate(structures):
        if np.char.str_len(name) < 1:
            structures[i] = structures[i-1]

    #replaces all instances of double underscore with single underscore
    for i, name in enumerate(structures):
        fixed_str = typo_correct(name, '_')
        structures[i] = fixed_str

    data_array[:, 0] = structures

    #Sometimes awkward shift where some data that should be in col 9 is in col 10.
    #Duplicate incorrect col10 into col 9 - data in cols10 onward isnt' useful anyways
    col10 = data_array[:,10]

    for i, item in enumerate(col10):
        if '<' in item or '>' in item:
            data_array[i, 9] = item
            data_array[i, 10] = data_array[i, 11]


    #Save as new xlsx file
    df = pd.DataFrame(data_array)
    df.to_excel(fpath, index=False)

    #Delete temporary csv file
    os.remove(output_path)
def find_criteria_index(dfn_criteria, desired_str):

    inds = []

    for idx, s in enumerate(dfn_criteria):
        if isinstance(s, str):
            if desired_str in s:
                inds.append(idx)
    inds = np.asarray(inds)

    return inds
def Gy_to_cGY(patient_dir, fraction_name):
    fpath_fraction = os.path.join(patient_dir, fraction_name + '.xlsx')
    df_fraction = pd.read_excel(fpath_fraction)

    #Determine units
    testcol = np.asarray(df_fraction.iloc[0:10,3])
    cGy_exist = False #Assume no cGy anywhere

    for i in testcol:
        if i == 'cGy':
            print('Gy to cGy conversion in process')
            cGy_exist = True #Change value to True only if at least one cell is cGy

    if cGy_exist == True:
        print('no need for conversion')
        return

    #convert dataframe to numpy array
    fraction_arr = df_fraction.values

    # Hard-code criteria dictionary
    value_dict_nospaces = {
        "V37Gy<=5cm3(+5cm3)": "V3700cGy <= 5 cm3 (+5 cm3)",
        "V18.1Gy<=40%": "V1810cGy <= 40 %",
        "V30Gy<=1cm3": "V3000cGy <= 1 cm3",
        "V18.1Gy<=5cm3": "V1810cGy <= 5 cm3",
        "V40Gy>=95%": "V4000cGy >= 95 %",
        "V36.25Gy>=95%": "V3625cGy >= 95 %",
        "D98%>=34.4Gy(-0.688Gy)": "D98% >= 3440 cGy (-0.688 Gy)",
        "V36Gy<=1cm3(+1cm3)": "V3600cGy <= 1 cm3 (+1 cm3)",
        "V29Gy<=20%": "V2900cGy <= 20 %",
        "V18.1Gy<=50%": "V1810cGy <= 50 %",
        "V42Gy<=50%(+50%)": "V4200cGy <= 50 % (+50 %)"
    }

    # Convert ATP rows from Gy to cGy if necessary
    if cGy_exist == False:
        cols_to_multiply = [3, 4, 5, 8]
        # Convert specified columns to integers, multiply by 100, and convert back to strings
        for i in range(len(fraction_arr)):
            try:
                fraction_arr[i, cols_to_multiply] = np.round(fraction_arr[i, cols_to_multiply].astype(float) * 100).astype(str)
            except:
                continue

        # Replace dose criteria strings
        arr_str = fraction_arr.astype(str) #Convert to string
        arr_str[:, 9] = np.char.replace(arr_str[:, 9], ' ', '')

        for key, value in value_dict_nospaces.items():
            print(key, value)
            exists = np.any(np.isin(arr_str, key))
            print(exists)
            arr_str = np.char.replace(arr_str, key, value)

        #Replace column unit labels
        arr_str[1, 3] = 'cGy'
        arr_str[1, 4] = 'cGy'
        arr_str[1, 5] = 'cGy'
        arr_str[1, 8] = 'cGy'

        #Overwrite dataframe with updated values
        df_fraction.iloc[0:, 0:] = arr_str
        print('Gy to cGy conversion - complete')

        # Overwrite original file with new dataframe
        df_fraction.to_excel(fpath_fraction, index=False, header=False)

        return
def concatenate_excel_files(patient_dir, ATP_name, fraction_name):

    fpath_ATP = os.path.join(patient_dir, ATP_name + '.xlsx')
    fpath_fraction = os.path.join(patient_dir, fraction_name + '.xlsx')

    #Read in excel files as dataframes
    df_main = pd.read_excel(fpath_fraction)
    df_ATP = pd.read_excel(fpath_ATP)

    #Find rows in df_ATP where Column0 contains '_ATP'
    filtered_rows = df_ATP[df_ATP.iloc[:,0].str.contains('_ATP')]

    #Check whether ATP stuff already appended to _Vx file. If so, don't proceed with code
    test = df_main[df_main.iloc[:,0].str.contains('_ATP')]
    test_arr = test.values
    if len(test_arr) > 0:
        print('ATP structures already in excel file')
        return


    #Convert dataframes to numpy arrays
    main_arr = df_main.values
    rows_arr = filtered_rows.values

    #Hard-code criteria dictionary
    value_dict = {
        "V37Gy <= 5 cm3 (+5 cm3)" : "V3700cGy <= 5 cm3 (+5 cm3)",
        "V18.1Gy <= 40 %" : "V1810cGy <= 40 %",
        "V30Gy <= 1 cm3" : "V3000cGy <= 1 cm3",
        "V18.1Gy <= 5 cm3" : "V1810cGy <= 5 cm3",
        "V40Gy >= 95 %" : "V4000cGy >= 95 %",
        "V36.25Gy >= 95 %" : "V3625cGy >= 95 %",
        "D98% >= 34.4 Gy (-0.688 Gy)": "D98% >= 3440 cGy (-0.688 Gy)",
        "V36Gy <= 1 cm3 (+1 cm3)": "V3600cGy <= 1 cm3 (+1 cm3)",
        "V29Gy <= 20 %": "V2900cGy <= 20 %",
        "V18.1Gy <= 50 %": "V1810cGy <= 50 %",
        "V42Gy <= 50 %" : "V4200cGy <= 50 %"
    }

    #Convert ATP rows from Gy to cGy if necessary
    val_unit = df_ATP.iloc[2, 3]
    if val_unit == 'Gy':
        cols_to_multiply = [3, 4, 5, 8]
        # Convert specified columns to integers, multiply by 10, and convert back to strings
        rows_arr[:, cols_to_multiply] = np.round(rows_arr[:, cols_to_multiply].astype(float) * 100).astype(str)
        print('Gy to cGy conversion - complete')

        arr_str = rows_arr.astype(str)
        #Replace dose criteria strings
        for key, value in value_dict.items():
            arr_str = np.char.replace(arr_str, key, value)
        rows_arr = arr_str

    new_arr2 = np.vstack((main_arr, rows_arr))

    #Convert concatenated array back into dataframe
    column_names = df_main.columns.tolist()
    df_new = pd.DataFrame(new_arr2, columns=column_names)

    #Overwrite original file with new dataframe, including appended ATP rows
    df_new.to_excel(fpath_fraction, index=False, header=False)

    return
############################################################################################
def populate_dose_stat_xlsx(patient_dir, fraction, ATS):
    #get patient name
    patient_name = os.path.split(patient_dir)[1]

    #Duplicate template file for specified patient
    fpath_template_file = '/Users/sblackledge/Documents/audit_evolutivePOTD/dose_stat_template.xlsx'
    fpath_output_file = os.path.join(patient_dir, patient_name + '_dose_stats.xlsx')
    TF = os.path.isfile(fpath_output_file)
    if not TF:
        shutil.copy(fpath_template_file, fpath_output_file)

    # Open file for writing
    workbook = openpyxl.load_workbook(fpath_output_file)
    #Delete sheet1 (automatically created by openpyxl - annoying and unnecessary)
    orig_sheetnames = workbook.sheetnames
    if 'Sheet1' in orig_sheetnames:
        std = workbook['Sheet1']
        workbook.remove(std)

    # Set active worksheet to dose criteria and extract thresholds
    active_sheet = workbook['dose criteria']
    firstRow = 1
    firstCol = 2
    nCols = 4
    nRows = 10

    allCells = np.array([[cell.value for cell in row] for row in active_sheet.iter_rows()])
    thresholds = allCells[(firstRow):(firstRow + nRows), (firstCol):(firstCol + nCols)]

    #Define sheetnames and row index where data from each fraction should be written
    sheetnames = ['Fraction 1', 'Fraction 2', 'Fraction 3', 'Fraction 4', 'Fraction 5']
    row_start_inds = [3, 17, 31, 45, 59, 73, 87]
    threshold_labels = ['Optimal', 'Mandatory', 'Marginal', 'Unacceptable']

    #Extract data from DVH stats excel file - convert to pandas dataframe
    if ATS == 1:
        fpath_dvh = os.path.join(patient_dir, patient_name + '_DVH stats MR' + str(fraction) + '.xlsx')
        suffixes = ['', '_V', '_CT', '_MR1', '_MR2', '_MR3', '_MR4']
        delivery_type = 'Delivery type: ATS'
    elif ATS == 0:  # ATP of ATS
        fpath_dvh = os.path.join(patient_dir, patient_name + '_DVH stats MR' + str(fraction) + '.xlsx')
        suffixes = ['', '_ATP', '_CT', '_MR1', '_MR2', '_MR3', '_MR4']
        delivery_type = 'Delivery type: ATP of ATS'
    elif ATS == 2:
        fpath_dvh = os.path.join(patient_dir, patient_name + '_DVH stats MR' + str(fraction) + '.xlsx')
        suffixes = ['', '_CT', '_MR1', '_MR2', '_MR3', '_MR4']
        delivery_type = 'Delivery type: unknown'

    df = pd.read_excel(fpath_dvh)
    dfn = df.to_numpy()
    #sf = 1

    dfn_structures = dfn[:,0]
    dfn_criteria = dfn[:, 9]
    dfn_criteria = list(dfn_criteria)

    # Set active worksheet for specified fraction
    active_sheet = workbook[sheetnames[fraction - 1]]
    print(active_sheet)

    #Correct structures names so all on one line (some are multiline by default)
    for i, s in enumerate(dfn_structures):
        is_float = isinstance(s, float)
        if not is_float:
            if "\n" in s:
                temp= s.splitlines()
                oneliner = "".join(temp)
                dfn_structures[i] = oneliner

    structure_basenames = ['CTVpsv_4000', 'PTVpsv_3625', 'Bladder', 'Rectum', 'Bowel']


    for f in range(fraction+2):
        structures = [structure_basename + suffixes[f] for structure_basename in structure_basenames]
        print(structures)

        #Initialize vectors to store DVHstats and score (whether optimal, mandatory, marginal, or unacceptable)
        stats = []

        #Extract dose stats for online plan (session)
        #CTV
        try:
            ind0 = np.where(dfn_structures == structures[0])[0][0]
            ctvstats = float(dfn[ind0, 7])
            stats.append(ctvstats)
        except:
            print(f"Structure '{structures[0]}' does not exist")

        #PTV
        inds_structures = np.where(dfn_structures == structures[1])[0] #rows containing requested structure
        if inds_structures.size == 0:
            print('No structure found in excel file matching ', structures[1])
        criteria = ['V3625', 'D98']
        col_inds = [7, 8]
        for i, c in enumerate(criteria):
            inds_criteria = find_criteria_index(dfn_criteria, c)
            try:
                ind = np.intersect1d(inds_structures, inds_criteria)[0]
                val = float(dfn[ind, col_inds[i]])
                stats.append(val)
            except:
                print(structures[1], ' does not have criteria ', c)
                val = float("nan")
                stats.append(val)

        #Bladder
        inds_structures = np.where(dfn_structures == structures[2])[0]
        if inds_structures.size == 0:
            print('No structure found in excel file matching ', structures[2])
        criteria = ['V3700', 'V1810']
        col_inds = [6, 7]
        for i, c in enumerate(criteria):
            inds_criteria = find_criteria_index(dfn_criteria, c)
            try:
                ind = np.intersect1d(inds_structures, inds_criteria)[0]
                val = float(dfn[ind, col_inds[i]])
                stats.append(val)
            except:
                print(structures[2], ' does not have criteria ', c)
                val = float("nan")
                stats.append(val)


        #Rectum
        inds_structures = np.where(dfn_structures == structures[3])[0]
        if inds_structures.size == 0:
            print('No structure found in excel file matching ', structures[3])
        criteria = ['V3600', 'V2900', 'V1810']
        col_inds = [6, 7, 7]
        for i, c in enumerate(criteria):
            inds_criteria = find_criteria_index(dfn_criteria, c)
            try:
                ind = np.intersect1d(inds_structures, inds_criteria)[0]
                val = float(dfn[ind, col_inds[i]])
                stats.append(val)
            except:
                print(structures[3], ' does not have criteria ', c)
                val = float("nan")
                stats.append(val)

        #Bowel
        inds_structures = np.where(dfn_structures == structures[4])[0]
        if inds_structures.size == 0:
            print('No structure found in excel file matching ', structures[4])
        criteria = ['V3000', 'V1810']
        col_inds = [6, 6]
        for i, c in enumerate(criteria):
            inds_criteria = find_criteria_index(dfn_criteria, c)
            try:
                ind = np.intersect1d(inds_structures, inds_criteria)[0]
                val = float(dfn[ind, col_inds[i]])
                stats.append(val)
            except:
                print(structures[4], ' does not have criteria ', c)
                val = float("nan")
                stats.append(val)

        #Adjust units
        #stats = [i * sf for i in stats]

        #Determine score for each dose stat
        score = []

        #CTV
        jj = 0
        if stats[jj] >= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        elif thresholds[jj, 1] <= stats[jj] < thresholds[jj, 0]:
            score_i = threshold_labels[1]
        elif thresholds[jj, 2] <= stats[jj] < thresholds[jj, 1]:
            score_i = threshold_labels[2]
        elif stats[jj] < thresholds[jj, 2]:
            score_i = threshold_labels[3]

        score.append(score_i)

        #Vol (%) PTVpsv_3625 covered by 36.25 Gy
        jj = 1
        if stats[jj] >= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        elif thresholds[jj, 1] <= stats[jj] < thresholds[jj, 0]:
            score_i = threshold_labels[1]
        elif thresholds[jj, 2] <= stats[jj] < thresholds[jj, 1]:
            score_i = threshold_labels[2]
        elif stats[jj] < thresholds[jj, 2]:
            score_i = threshold_labels[3]

        score.append(score_i)

        #Dose (Gy) covering 98% of PTVpsv_3625
        jj = 2
        if stats[jj] >= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        elif thresholds[jj, 1] <= stats[2] < thresholds[jj, 0]:
            score_i = threshold_labels[1]
        elif stats[jj] < thresholds[jj, 1]:
            score_i = threshold_labels[3]

        score.append(score_i)

        #Volume(cc) of bladder covered by 37Gy
        jj = 3
        if stats[jj] <= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        elif thresholds[jj, 0] < stats[jj] <= thresholds[jj, 1]:
            score_i = threshold_labels[1]
        elif stats[jj] > thresholds[jj, 1]:
            score_i = threshold_labels[3]

        score.append(score_i)

        #Volume (%) of bladder covered by 18.10Gy
        jj = 4
        if stats[jj] <= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        else:
            score_i = threshold_labels[3]

        score.append(score_i)

        #Volume (cc) of rectum covered by 36 Gy
        jj = 5
        if stats[jj] <= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        elif thresholds[jj, 0] < stats[jj] <= thresholds[jj, 1]:
            score_i = threshold_labels[1]
        elif stats[jj] > thresholds[jj, 1]:
            score_i = threshold_labels[3]

        score.append(score_i)

        #Volume(%) of Rectum covered by 29 Gy
        jj = 6
        if stats[jj] <= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        else:
            score_i = threshold_labels[3]

        score.append(score_i)

        #Volume(%) of Rectum covered by 18.10 Gy
        jj = 7
        if stats[jj] <= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        else:
            score_i = threshold_labels[3]

        score.append(score_i)

        #Volume (cc) of Bowel covered by 30 Gy
        jj = 8
        if stats[jj] <= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        else:
            score_i = threshold_labels[3]

        score.append(score_i)

        #Volume (cc) of Bowel coveredy by 18.10 gy
        jj = 9
        if stats[jj] <= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        else:
            score_i = threshold_labels[3]

        score.append(score_i)

        #Write stats vector to excel file
        row_start = row_start_inds[f]
        for ii, j in enumerate(stats):
            val = round(row_start + ii)
            active_sheet.cell(row=val, column=3).value = j
            active_sheet.cell(row=val, column=4).value = score[ii]

        #Write whether ATS or ATP of ATS plan
        active_sheet.cell(row=1, column=2).value = delivery_type

        workbook.save(fpath_output_file)
##################################################################################
def extract_dose_stats(structures, dfn_structures, structure_id, dfn_criteria, dfn):
    structure_dict = {
        #CTV
        0: [
            ['V4000'],
            [7]
        ],
        #PTV
        1: [
            ['V3625', 'D98'],
            [7, 8]
        ],
        #Bladder
        2: [
            ['V3700', 'V1810'],
            [6, 7],
        ],
        #Rectum
        3: [
            ['V3600', 'V2900', 'V1810'],
            [6, 7, 7]
        ],
        #Bowel
        4: [
            ['V3000', 'V1810'],
            [6, 6]
        ],
        #Urethra
        5: [
            ['V4200'],
            [7]

        ]
    }
    s = []
    structure = structures[structure_id]
    criteria = structure_dict[structure_id][0]
    col_inds = structure_dict[structure_id][1]
    inds_structures = np.where(dfn_structures == structure)[0]
    if inds_structures.size == 0:
        print('No structure found in excel file matching ', structure)
    for i, c in enumerate(criteria):
        inds_criteria = find_criteria_index(dfn_criteria, c)
        try:
            ind = np.intersect1d(inds_structures, inds_criteria)[0]
            val = float(dfn[ind, col_inds[i]])
            s.append(val)
        except:
            print(structure, ' does not have criteria ', c)
            val = float("nan")
            s.append(val)

    return s

def populate_dose_stat_ROInew_xlsx(patient_dir, fraction, ATS):
    #get patient name
    patient_name = os.path.split(patient_dir)[1]

    #Duplicate template file for specified patient
    fpath_template_file = '/Users/sblackledge/Documents/audit_evolutivePOTD/dose_stat_template.xlsx'
    fpath_output_file = os.path.join(patient_dir, patient_name + '_dose_stats.xlsx')
    TF = os.path.isfile(fpath_output_file)
    if not TF:
        shutil.copy(fpath_template_file, fpath_output_file)

    # Open file for writing
    workbook = openpyxl.load_workbook(fpath_output_file)
    #Delete sheet1 (automatically created by openpyxl - annoying and unnecessary)
    orig_sheetnames = workbook.sheetnames
    if 'Sheet1' in orig_sheetnames:
        std = workbook['Sheet1']
        workbook.remove(std)

    # Set active worksheet to dose criteria and extract thresholds
    active_sheet = workbook['dose criteria']
    firstRow = 1
    firstCol = 2
    nCols = 4
    nRows = 11

    allCells = np.array([[cell.value for cell in row] for row in active_sheet.iter_rows()])
    thresholds = allCells[(firstRow):(firstRow + nRows), (firstCol):(firstCol + nCols)]

    #Define sheetnames and row index where data from each fraction should be written
    threshold_labels = ['Optimal', 'Mandatory', 'Marginal', 'Unacceptable']
    row_start_inds = [33, 48, 63, 78, 93] #backwards propagations
    sheetnames = ['Fraction 1', 'Fraction 2', 'Fraction 3', 'Fraction 4', 'Fraction 5']

    #Extract data from DVH stats excel file - convert to pandas dataframe
    if fraction == 0:
        fpath_dvh = os.path.join(patient_dir, patient_name + '_DVH stats CT' + '.xlsx')
        suffixes = ['_MR1', '_MR2', '_MR3', '_MR4', '_MR5']
        sheetnames = ['Fraction 1', 'Fraction 2', 'Fraction 3', 'Fraction 4', 'Fraction 5']

    else:
        fpath_dvh = os.path.join(patient_dir, patient_name + '_DVH stats MR' + str(fraction) + '.xlsx')
        sheet_str = f'Fraction {fraction}'
        #Append sheetnames with fraction id twice to write session and verif dose stats
        sheetnames.append(sheet_str)
        sheetnames.append(sheet_str)

        if ATS == 0:
            suffixes = ['_MR1', '_MR2', '_MR3', '_MR4', '_MR5', '', '_ATP']
        elif ATS == 1:
            suffixes = ['_MR1', '_MR2', '_MR3', '_MR4', '_MR5', '', '_V']


    df = pd.read_excel(fpath_dvh)
    dfn = df.to_numpy()
    dfn_structures = dfn[:,0]
    dfn_criteria = dfn[:, 9]
    dfn_criteria = list(dfn_criteria)

    #Correct structures names so all on one line (some are multiline by default)
    for i, s in enumerate(dfn_structures):
        is_float = isinstance(s, float)
        if not is_float:
            if "\n" in s:
                temp= s.splitlines()
                oneliner = "".join(temp)
                dfn_structures[i] = oneliner


    structure_basenames = ['CTVpsv_4000', 'PTVpsv_3625', 'Bladder', 'Rectum', 'Bowel', 'Urethra']

    for fr in range(fraction, len(sheetnames)):
    #for i, f in enumerate(suffixes):
        active_sheet = workbook[sheetnames[fr]]
        print(active_sheet)
        f = suffixes[fr]
        print(fr)


        structures = [structure_basename + f for structure_basename in structure_basenames]

        #Initialize vectors to store DVHstats and score (whether optimal, mandatory, marginal, or unacceptable)
        stats = []

        #Extract dose stats
        for k in range(6):
            s = extract_dose_stats(structures, dfn_structures, k, dfn_criteria, dfn)
            stats.append(s)

        #flatten
        stats = [item for sublist in stats for item in (sublist if isinstance(sublist, list) else [sublist])]

        #Determine score for each dose stat
        score = []

        #CTV
        jj = 0
        if stats[jj] >= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        elif thresholds[jj, 1] <= stats[jj] < thresholds[jj, 0]:
            score_i = threshold_labels[1]
        elif thresholds[jj, 2] <= stats[jj] < thresholds[jj, 1]:
            score_i = threshold_labels[2]
        elif stats[jj] < thresholds[jj, 2]:
            score_i = threshold_labels[3]

        score.append(score_i)

        #Vol (%) PTVpsv_3625 covered by 36.25 Gy
        jj = 1
        if stats[jj] >= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        elif thresholds[jj, 1] <= stats[jj] < thresholds[jj, 0]:
            score_i = threshold_labels[1]
        elif thresholds[jj, 2] <= stats[jj] < thresholds[jj, 1]:
            score_i = threshold_labels[2]
        elif stats[jj] < thresholds[jj, 2]:
            score_i = threshold_labels[3]

        score.append(score_i)

        #Dose (Gy) covering 98% of PTVpsv_3625
        jj = 2
        if stats[jj] >= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        elif thresholds[jj, 1] <= stats[2] < thresholds[jj, 0]:
            score_i = threshold_labels[1]
        elif stats[jj] < thresholds[jj, 1]:
            score_i = threshold_labels[3]

        score.append(score_i)

        #Volume(cc) of bladder covered by 37Gy
        jj = 3
        if stats[jj] <= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        elif thresholds[jj, 0] < stats[jj] <= thresholds[jj, 1]:
            score_i = threshold_labels[1]
        elif stats[jj] > thresholds[jj, 1]:
            score_i = threshold_labels[3]

        score.append(score_i)

        #Volume (%) of bladder covered by 18.10Gy
        jj = 4
        if stats[jj] <= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        else:
            score_i = threshold_labels[3]

        score.append(score_i)

        #Volume (cc) of rectum covered by 36 Gy
        jj = 5
        if stats[jj] <= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        elif thresholds[jj, 0] < stats[jj] <= thresholds[jj, 1]:
            score_i = threshold_labels[1]
        elif stats[jj] > thresholds[jj, 1]:
            score_i = threshold_labels[3]

        score.append(score_i)

        #Volume(%) of Rectum covered by 29 Gy
        jj = 6
        if stats[jj] <= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        else:
            score_i = threshold_labels[3]

        score.append(score_i)

        #Volume(%) of Rectum covered by 18.10 Gy
        jj = 7
        if stats[jj] <= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        else:
            score_i = threshold_labels[3]

        score.append(score_i)

        #Volume (cc) of Bowel covered by 30 Gy
        jj = 8
        if stats[jj] <= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        else:
            score_i = threshold_labels[3]

        score.append(score_i)

        #Volume (cc) of Bowel coveredy by 18.10 gy
        jj = 9
        if stats[jj] <= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        else:
            score_i = threshold_labels[3]

        score.append(score_i)

        # Urethra
        jj = 10
        if stats[jj] <= thresholds[jj, 0]:
            score_i = threshold_labels[0]
        else:
            score_i = threshold_labels[1]

        score.append(score_i)

        #Write stats vector to excel file
        if fr <= 4:
            row_start_ind = row_start_inds[fraction]
        elif fr == 5:
            row_start_ind = 3
        elif fr == 6:
            row_start_ind = 18

        for ii, j in enumerate(stats):
            val = round(row_start_ind + ii)
            active_sheet.cell(row=val, column=3).value = j
            active_sheet.cell(row=val, column=4).value = score[ii]

        #Write whether ATS or ATP of ATS plan
        #active_sheet.cell(row=1, column=2).value = delivery_type

        workbook.save(fpath_output_file)

###################################################################################
#Specific fraction
patient_name = 'PER093'
patient_dir = os.path.join('/Users/sblackledge/Documents/audit_evolutivePOTD/ROInew', patient_name)
#fraction = 0 for ref CT, 1 for MR1, 2 for MR2, etc.
fraction = 0
#ATS: 1 if ATS used, 0 if ATP of ATS used. 2 otherwise. Check 'ATP' column in 'ROI propagation key.xlsx'
ATS = 1

#If assessing suitability of reference plan, ATP/ATS is irrelevant - just need to plop ROIs onto plan. Verif doesn't exist
if fraction == 0:
    fraction_name = patient_name + '_DVH stats CT'
    pdf_to_csv(fraction_name, patient_dir)
    Gy_to_cGY(patient_dir, fraction_name)

#Here, ATP/ATS is important for determining dose actually delivered on verif.
else:
    if ATS == 1: #ATS
        fraction_name = patient_name + '_DVH stats MR' + str(fraction)
        pdf_to_csv(fraction_name, patient_dir)
        Gy_to_cGY(patient_dir, fraction_name)

    elif ATS == 0:  # ATP of ATS
        fraction_name = patient_name + '_DVH stats MR' + str(fraction)
        ATP_name = patient_name + '_DVH stats MR' + str(fraction) + '_ATP'
        pdf_to_csv(fraction_name, patient_dir)
        pdf_to_csv(ATP_name, patient_dir)
        Gy_to_cGY(patient_dir, fraction_name)
        # Append ATP structures (dose stats from verif contours propagated to delivered plan) to fraction_name excel file
        concatenate_excel_files(patient_dir, ATP_name, fraction_name)

fractions = [0, 1, 2, 3, 4, 5]
delivery_type = [1, 1, 1, 1, 0, 0]
for i, f in enumerate(fractions):
    populate_dose_stat_ROInew_xlsx(patient_dir, f, delivery_type[i])

populate_dose_stat_ROInew_xlsx(patient_dir, 0, 1)